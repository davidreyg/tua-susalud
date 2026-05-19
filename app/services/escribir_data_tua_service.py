"""Servicio para escribir datos TUA en la plantilla Excel."""

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.api.responses.tua_input_response import TuaInputDataResponse
from tools.logger import Logger

logger = Logger(__name__)

_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "utils" / "TUASUSALUD.xlsx"

# Columnas de la hoja "Data" (1-indexed)
_COL_CODIGO_UNICO = 1  # A
_COL_NOMBRE = 2  # B
_COL_RED = 3  # C
_COL_TIPO_DOC = 4  # D
_COL_NUM_DOC = 5  # E
_COL_PROFESION = 6  # F
_COL_NUM_COLEG = 7  # G
_COL_APELLIDOS = 8  # H
_COL_NOMBRES = 9  # I
_COL_ESPECIALIDAD = 10  # J
_COL_SERVICIO = 11  # K
_COL_INGRESO_BASE = 12  # L (Día 1 Ingreso)
_COL_SALIDA_BASE = 13  # M (Día 1 Salida)

_FILA_INICIO_DATOS = 4
_MAX_FILAS = 10000
_MAX_DIAS = 31
_MAX_COLUMNAS = 73  # Hasta BU


class EscribirDataTuaService:
    """Servicio para escribir datos TUA en la plantilla Excel.

    Carga la plantilla ``TUASUSALUD.xlsx``, escribe los registros
    en la hoja ``Data`` respetando la estructura de columnas
    predefinida y retorna el archivo como bytes.
    """

    @staticmethod
    def escribir(datos: list[TuaInputDataResponse]) -> bytes:
        """Escribe los registros TUA en la plantilla y retorna el Excel.

        Args:
            datos: Lista de registros TUA a escribir.

        Returns:
            Bytes del archivo Excel generado.

        Raises:
            ValueError: Si la lista está vacía o excede la capacidad
                de la plantilla.

        """
        logger.info(
            "Iniciando escritura de %d registro(s) en la plantilla...",
            len(datos),
        )

        if not datos:
            msg = "No hay datos para escribir."
            logger.warning(msg)
            raise ValueError(msg)

        max_registros = _MAX_FILAS - _FILA_INICIO_DATOS + 1
        if len(datos) > max_registros:
            logger.error(
                "Cantidad de registros (%d) excede el máximo (%d).",
                len(datos),
                max_registros,
            )
            msg = (
                f"La cantidad de registros ({len(datos)}) excede el "
                f"máximo permitido ({max_registros})."
            )
            raise ValueError(msg)

        if not _TEMPLATE_PATH.exists():
            logger.error("Plantilla no encontrada: %s", _TEMPLATE_PATH)
            msg = f"Plantilla no encontrada: {_TEMPLATE_PATH}"
            raise FileNotFoundError(msg)

        wb: Workbook = load_workbook(str(_TEMPLATE_PATH))
        try:
            ws = wb["Data"]

            EscribirDataTuaService._limpiar_datos_existentes(ws)
            EscribirDataTuaService._escribir_registros(ws, datos)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            result = output.getvalue()

            logger.info(
                "Escritura completada (%d registros, %d bytes).",
                len(datos),
                len(result),
            )
        except Exception:
            logger.exception("Error al escribir en la plantilla Excel.")
            raise
        else:
            return result
        finally:
            del wb

    @staticmethod
    def _limpiar_datos_existentes(ws: Worksheet) -> None:
        """Limpia los datos existentes en el área de datos de la hoja.

        Args:
            ws: Hoja de trabajo a limpiar.

        """
        for row in ws.iter_rows(
            min_row=_FILA_INICIO_DATOS,
            max_row=_MAX_FILAS,
            max_col=_MAX_COLUMNAS,
        ):
            for cell in row:
                cell.value = None

        logger.debug(
            "Datos existentes limpiados (filas %d-%d).",
            _FILA_INICIO_DATOS,
            _MAX_FILAS,
        )

    @staticmethod
    def _escribir_registros(
        ws: Worksheet,
        datos: list[TuaInputDataResponse],
    ) -> None:
        """Escribe los registros en la hoja a partir de la fila de inicio.

        Args:
            ws: Hoja de trabajo donde escribir.
            datos: Lista de registros TUA.

        """
        for i, record in enumerate(datos):
            row_num = _FILA_INICIO_DATOS + i

            ws.cell(row=row_num, column=_COL_CODIGO_UNICO, value=record.codigo_unico)
            _nombre = record.nombre_establecimiento
            ws.cell(row=row_num, column=_COL_NOMBRE, value=_nombre)
            ws.cell(row=row_num, column=_COL_RED, value=record.red)
            ws.cell(row=row_num, column=_COL_TIPO_DOC, value=record.tipo_documento)
            ws.cell(row=row_num, column=_COL_NUM_DOC, value=record.numero_documento)
            ws.cell(row=row_num, column=_COL_PROFESION, value=record.profesion)
            ws.cell(row=row_num, column=_COL_NUM_COLEG, value=record.numero_colegiatura)
            ws.cell(row=row_num, column=_COL_APELLIDOS, value=record.apellidos)
            ws.cell(row=row_num, column=_COL_NOMBRES, value=record.nombres)
            ws.cell(row=row_num, column=_COL_ESPECIALIDAD, value=record.especialidad)
            ws.cell(row=row_num, column=_COL_SERVICIO, value=record.servicio)

            EscribirDataTuaService._escribir_turnos(ws, row_num, record.turnos)

    @staticmethod
    def _escribir_turnos(
        ws: Worksheet,
        row_num: int,
        turnos: list,
    ) -> None:
        """Escribe los turnos de un registro en las columnas correspondientes.

        Args:
            ws: Hoja de trabajo.
            row_num: Número de fila del registro.
            turnos: Lista de objetos TurnoItem.

        """
        turnos_map = {t.dia: t for t in turnos if t.dia is not None}

        for dia in range(1, _MAX_DIAS + 1):
            col_ingreso = _COL_INGRESO_BASE + (dia - 1) * 2
            col_salida = _COL_SALIDA_BASE + (dia - 1) * 2

            turno = turnos_map.get(dia)
            if turno:
                if turno.hora_entrada:
                    ws.cell(row=row_num, column=col_ingreso, value=turno.hora_entrada)
                if turno.hora_salida:
                    ws.cell(row=row_num, column=col_salida, value=turno.hora_salida)
